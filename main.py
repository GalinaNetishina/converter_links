import os
import webbrowser

import requests

from flask import Flask, flash, request, redirect, url_for, render_template, send_file

from openpyxl import load_workbook

import os
from dotenv import load_dotenv
dotenv_path = os.path.join(os.path.dirname(__file__), '.env')
if os.path.exists(dotenv_path):
    load_dotenv(dotenv_path)
def get_token():
    from urllib.parse import urlencode
    APP_ID ="51800837"
    base_url = "https://oauth.vk.com/authorize"
    params = {
        "client_id": APP_ID,
        "redirect_uri": "https://oauth.vk.com/blank.html",
        "display": "popup",
        "scope": "status",
        "response_type": "token"
    }
    url = f"{base_url}?{urlencode(params)}"
    webbrowser.open(url, params=params)



app = Flask(__name__)

VK_TOKEN = os.getenv("VK_TOKEN") or get_token()


def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'xlsx'}
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/', methods=['GET', 'POST'])
async def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Не могу прочитать файл')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('Нет выбранного файла')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            file.save('source.xlsx')
            await _load()
            return send_file('result.xlsx', as_attachment=True)
    return render_template("index.html")


async def _load():
    wb = load_workbook('source.xlsx')
    ws = wb[wb.sheetnames[0]]
    for i, link in enumerate((cell.value for cell in ws['A']), start=1):
        ws[f"B{i}"].value = get_short(link)
    wb.save('result.xlsx')


def get_short(long_link):
    url = 'https://api.vk.com/method/utils.getShortLink'
    params = {'url': long_link,
              'v': "5.199",
              'access_token': VK_TOKEN
              }
    response = requests.get(url, params=params)
    if response.ok:
        return response.json()["response"]["short_url"]
    return 'error'

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=3000)